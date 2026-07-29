import hashlib
from django.utils import timezone
from django.http import HttpResponse
from rest_framework import viewsets, permissions, status, generics
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import BasePermission, AllowAny
from rest_framework.views import APIView

from .pdf_reports import generate_project_alignment_report
from .models import (
    Project, Sector, SDG, SDGTarget, SDGIndicator,
    ProjectSDGAlignment, SND30Axis, ProjectSND30Alignment,
    ProjectWorkflowTrace, OddRoleRequest, Preuve,
    Period, MesureKPI, CampagneReporting, Rapport,
    PCDObjective,
    GaddDimension, GaddObjective, GaddEvaluation, GaddObjectiveAnswer,
)
from .serializers import (
    ProjectSerializer, SectorSerializer, SDGSerializer, SND30AxisSerializer,
    ProjectSDGAlignmentSerializer, ProjectSND30AlignmentSerializer,
    WorkflowTraceSerializer, OddRoleRequestSerializer, PreuveSerializer,
    PeriodSerializer, MesureKPISerializer,
    CampagneReportingSerializer, RapportSerializer,
    PCDObjectiveSerializer,
    GaddDimensionSerializer, GaddEvaluationSerializer,
)

class ProjectPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100

class PublicProjectViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet pour la consultation publique des projets certifiés."""
    serializer_class = ProjectSerializer
    permission_classes = [AllowAny]
    pagination_class = ProjectPagination

    def get_queryset(self):
        return Project.objects.filter(is_published=True).order_by('-published_at')

class ProjectExportPDFView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        try:
            project = Project.objects.get(pk=pk)
        except Project.DoesNotExist:
            return Response({'detail': 'Projet non trouvé.'}, status=status.HTTP_404_NOT_FOUND)
        
        # On recalcule les scores avant l'export pour être sûr qu'ils sont à jour
        project.update_scores()
        
        pdf_content = generate_project_alignment_report(project)
        
        response = HttpResponse(pdf_content, content_type='application/pdf')
        filename = f"Fiche_Alignement_ODD_{project.id}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

# ── RÔLES SUSPENDUS — tout est ouvert à tous les utilisateurs authentifiés ──
ROLES_ENABLED = True

# ── Noms canoniques des rôles ODD ──────────────────────────────────────────
ROLE_SUPER_ADMIN        = 'Super Admin'
ROLE_ADMIN_ARRDEL       = 'Admin ARRDEL'
ROLE_ADMIN_CTD          = 'Admin CTD'
ROLE_POINT_FOCAL        = 'Point Focal ODD'
ROLE_RESPONSABLE_PROJET = 'Responsable Projet'
ROLE_VALIDATEUR_CTD     = 'Validateur CTD'
ROLE_VALIDATEUR_ARRDEL  = 'Validateur ARRDEL'

ADMIN_ROLES     = {ROLE_SUPER_ADMIN, ROLE_ADMIN_ARRDEL, ROLE_ADMIN_CTD}
VALIDATOR_ROLES = {ROLE_VALIDATEUR_CTD, ROLE_VALIDATEUR_ARRDEL}

# Transitions autorisées par rôle.
# La soumission place le projet en 'submitted' et l'y laisse : la vérification
# de complétude (auto_validate(), cf. check_completeness ci-dessous) est
# déclenchée à la demande par un Validateur CTD/ARRDEL ou un Administrateur,
# pas automatiquement à l'instant de la soumission.
# Les rôles humains gèrent ensuite les transitions manuelles autorisées.
ALLOWED_TRANSITIONS = {
    ROLE_RESPONSABLE_PROJET: {
        'draft':            ['submitted'],
        'rejected':         ['draft'],
        'needs_completion': ['draft'],
    },
    ROLE_ADMIN_CTD: {
        'draft':            ['submitted'],
        'rejected':         ['draft'],
        'needs_completion': ['draft'],
    },
    ROLE_VALIDATEUR_CTD: {
        'submitted':          ['controlling_ctd', 'validated_ctd', 'rejected', 'needs_completion'],
        'controlling_ctd':    ['validated_ctd', 'rejected', 'needs_completion'],
    },
    ROLE_VALIDATEUR_ARRDEL: {
        'validated_ctd':      ['controlling_arrdel', 'validated', 'rejected', 'needs_completion'],
        'controlling_arrdel': ['validated', 'rejected', 'needs_completion'],
    },
}


def user_roles(user) -> set:
    if not ROLES_ENABLED:
        return set()
    if not getattr(user, 'is_authenticated', False):
        return set()
    return set(user.roles.values_list('name', flat=True))


def is_admin(user) -> bool:
    if not ROLES_ENABLED:
        return True
    return user.is_staff or user.is_superuser or bool(user_roles(user) & ADMIN_ROLES)


def roles_configured(user) -> bool:
    if not ROLES_ENABLED:
        return False
    return bool(user_roles(user))


def can_transition(roles: set, current: str, target: str, is_owner: bool = False) -> bool:
    if not ROLES_ENABLED:
        return True
    if not roles:
        return True
    if roles & ADMIN_ROLES:
        return True

    if target in ['validated', 'validated_ctd', 'controlling_ctd', 'controlling_arrdel'] and not (roles & VALIDATOR_ROLES):
        return False

    for role in roles:
        allowed = ALLOWED_TRANSITIONS.get(role, {}).get(current, [])
        if target in allowed:
            if target in ['draft', 'submitted'] and not is_owner and not (roles & ADMIN_ROLES):
                continue
            return True
    return False


def notify_status_change(project: Project, new_status: str) -> None:
    """Envoie une notification email (silencieuse en cas d'erreur)."""
    try:
        from django.core.mail import send_mail
        from django.conf import settings as conf

        status_labels = {
            'submitted':          'Soumis pour validation',
            'controlling_ctd':    'Pris en charge par CTD',
            'validated_ctd':      'Validé CTD',
            'controlling_arrdel': 'Pris en charge par ARRDEL',
            'validated':          'Aligné et certifié',
            'rejected':           'Rejeté',
        }
        label = status_labels.get(new_status, new_status)
        
        recipients = set([project.owner.email])
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        if new_status == 'submitted':
            ctd_validators = User.objects.filter(roles__name=ROLE_VALIDATEUR_CTD, is_active=True).values_list('email', flat=True)
            recipients.update(ctd_validators)
        elif new_status == 'validated_ctd':
            arrdel_validators = User.objects.filter(roles__name=ROLE_VALIDATEUR_ARRDEL, is_active=True).values_list('email', flat=True)
            recipients.update(arrdel_validators)

        subject = f"[ArrdelBee] Projet « {project.name} » — {label}"
        body = (
            f"Bonjour,\n\n"
            f"Le statut du projet « {project.name} » a été mis à jour : {label}.\n"
        )
        if new_status == 'rejected' and project.rejection_reason:
            body += f"\nMotif : {project.rejection_reason}\n"
        body += "\nConnectez-vous sur ArrdelBee pour voir les détails.\n"

        from_email = getattr(conf, 'DEFAULT_FROM_EMAIL', 'no-reply@arrdelbee.cm')
        send_mail(subject, body, from_email, list(recipients), fail_silently=True)
    except Exception:
        pass  # Ne jamais bloquer une action pour une notification


# ──────────────────────────────────────────────────────────────────────────
# ViewSets
# ──────────────────────────────────────────────────────────────────────────

class SectorViewSet(viewsets.ReadOnlyModelViewSet):
    """Référentiel des secteurs — lecture seule pour tous les utilisateurs connectés."""
    queryset = Sector.objects.all().order_by('name')
    serializer_class = SectorSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None


class ProjectViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = ProjectPagination

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return Project.objects.none()

        if not ROLES_ENABLED:
            return Project.objects.select_related('sector', 'commune').all().order_by('-created_at')

        user = self.request.user
        roles = user_roles(user)

        if user.is_staff or user.is_superuser or roles & ADMIN_ROLES:
            return Project.objects.select_related('sector', 'commune').all().order_by('-created_at')

        if not roles:
            # Aucun rôle assigné : accès non restreint, cohérent avec le reste
            # du module (transitions, actions) qui traite l'absence de rôle
            # comme "pas encore configuré" plutôt que "aucun droit".
            return Project.objects.select_related('sector', 'commune').all().order_by('-created_at')

        is_ctd    = ROLE_VALIDATEUR_CTD in roles
        is_arrdel = ROLE_VALIDATEUR_ARRDEL in roles

        if is_ctd and not is_arrdel:
            # Palier CTD : dossiers en attente de son contrôle + ce qu'il a déjà
            # traité (visibilité sur son propre historique, y compris une fois
            # certifiés par l'ARRDEL — sinon le dossier disparaît de sa vue dès
            # que l'ARRDEL termine), borné à sa commune quand elle est renseignée.
            qs = Project.objects.select_related('sector', 'commune').filter(
                status__in=['submitted', 'controlling_ctd', 'validated_ctd',
                            'validated', 'rejected', 'needs_completion']
            )
            if user.commune_id:
                qs = qs.filter(commune_id=user.commune_id)
            return qs.order_by('-created_at')

        if is_arrdel and not is_ctd:
            # Palier ARRDEL : uniquement ce qui a déjà passé le contrôle CTD —
            # certification finale, portée nationale (pas de filtre territoire).
            return Project.objects.select_related('sector', 'commune').filter(
                status__in=['validated_ctd', 'controlling_arrdel', 'validated',
                            'rejected', 'needs_completion']
            ).order_by('-created_at')

        if roles & VALIDATOR_ROLES:
            # Cumul des deux rôles (rare) : vue complète des deux paliers.
            return Project.objects.select_related('sector', 'commune').filter(
                status__in=['submitted', 'controlling_ctd', 'validated_ctd',
                            'controlling_arrdel', 'validated', 'rejected', 'needs_completion']
            ).order_by('-created_at')

        return Project.objects.select_related('sector', 'commune').filter(
            owner=user
        ).order_by('-created_at')

    @action(detail=False, methods=['get'], url_path='pending-validation-count')
    def pending_validation_count(self, request):
        # `tier` : utilisé par la page d'accueil de validation pour afficher
        # le compteur propre à chaque palier (CTD / ARRDEL), indépendamment
        # du rôle du visiteur. Sans ce paramètre, on retombe sur le queryset
        # scopé de l'utilisateur courant (badge de l'onglet Validation).
        tier = request.query_params.get('tier')
        if tier == 'ctd':
            count = Project.objects.filter(status__in=['submitted', 'controlling_ctd']).count()
        elif tier == 'arrdel':
            count = Project.objects.filter(status__in=['validated_ctd', 'controlling_arrdel']).count()
        else:
            count = self.get_queryset().filter(
                status__in=['submitted', 'controlling_ctd', 'controlling_arrdel']
            ).count()
        return Response({'count': count})

    @action(detail=True, methods=['post'], url_path='notify-arrdel')
    def notify_arrdel(self, request, pk=None):
        """
        Transmission explicite au Valideur ARRDEL une fois le contrôle CTD
        terminé — déclenche la notification (email) même si celle-ci a déjà
        pu partir automatiquement lors du passage à 'validated_ctd'.
        """
        roles = user_roles(request.user)
        if roles and not (roles & (ADMIN_ROLES | {ROLE_VALIDATEUR_CTD})):
            return Response(
                {'detail': 'Seul un Validateur CTD ou Administrateur peut transmettre au Valideur ARRDEL.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        project = self.get_object()
        if project.status != 'validated_ctd':
            return Response(
                {'detail': "Ce projet doit d'abord être validé au niveau CTD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        notify_status_change(project, 'validated_ctd')
        return Response({'detail': 'Le Valideur ARRDEL a été notifié.'})

    def create(self, request, *args, **kwargs):
        roles = user_roles(request.user)
        # N'applique la restriction que si des rôles sont configurés
        if roles and not (roles & (ADMIN_ROLES | {ROLE_POINT_FOCAL, ROLE_RESPONSABLE_PROJET})):
            return Response(
                {'detail': 'Seul un Responsable Projet, Point Focal ODD ou Administrateur peut créer un projet.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().create(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        project = self.get_object()
        # N'applique la restriction que si des rôles sont configurés
        if not is_admin(request.user) and project.owner != request.user:
            return Response(
                {'detail': 'Vous ne pouvez supprimer que vos propres projets ou être administrateur.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().destroy(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        project = self.get_object()
        new_status = request.data.get('status')

        if new_status and new_status != project.status:
            roles = user_roles(request.user)

            is_owner = project.owner == request.user
            if not can_transition(roles, project.status, new_status, is_owner=is_owner):
                return Response(
                    {'detail': f"Votre rôle ({', '.join(roles)}) ne permet pas la transition '{project.status}' → '{new_status}' pour ce projet."},
                    status=status.HTTP_403_FORBIDDEN,
                )

            if new_status == 'rejected':
                reason = request.data.get('rejection_reason', '').strip()
                if not reason:
                    return Response(
                        {'detail': 'Le motif de rejet est obligatoire.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            ProjectWorkflowTrace.objects.create(
                project=project,
                actor=request.user,
                from_status=project.status,
                to_status=new_status,
                comment=request.data.get('rejection_reason', '')
                        or request.data.get('comment', ''),
            )

            if new_status == 'rejected' and request.data.get('rejection_reason'):
                project.rejection_reason = request.data['rejection_reason']
                project.save(update_fields=['rejection_reason'])

        kwargs['partial'] = True
        result = self.update(request, *args, **kwargs)

        # Notifier par email après la mise à jour réussie
        if new_status and new_status != (kwargs.get('_prev_status') or ''):
            project.refresh_from_db()
            notify_status_change(project, project.status)

        return result

    @action(detail=True, methods=['post'], url_path='check-completeness')
    def check_completeness(self, request, pk=None):
        """
        Déclenche manuellement le moteur de vérification de complétude
        (odd/auto_validation.py) sur un projet en attente de revue.
        Réservé aux validateurs CTD/ARRDEL et administrateurs — remplace
        l'ancien déclenchement automatique et silencieux à la soumission.
        """
        project = self.get_object()
        roles = user_roles(request.user)
        if roles and not (roles & (ADMIN_ROLES | VALIDATOR_ROLES)):
            return Response(
                {'detail': 'Seul un Validateur CTD/ARRDEL ou Administrateur peut lancer cette vérification.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        if project.status not in ['submitted', 'controlling_ctd', 'needs_completion']:
            return Response(
                {'detail': f"La vérification de complétude ne s'applique pas au statut '{project.status}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from .auto_validation import auto_validate
        final_status, rejection_reason = auto_validate(project, actor=request.user)
        project.status = final_status
        project.rejection_reason = rejection_reason
        project.save(update_fields=['status', 'rejection_reason'])
        notify_status_change(project, final_status)
        serializer = self.get_serializer(project)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def sync_alignments(self, request, pk=None):
        roles = user_roles(request.user)
        if roles and not (roles & (ADMIN_ROLES | {ROLE_POINT_FOCAL})):
            return Response(
                {'detail': "Seul un Point Focal ODD ou Administrateur peut effectuer l'alignement."},
                status=status.HTTP_403_FORBIDDEN,
            )

        project = self.get_object()
        sdg_indicator_codes = request.data.get('sdg_indicators')
        snd30_axes_numbers  = request.data.get('snd30_axes')
        local_priorities    = request.data.get('local_priorities')
        force               = bool(request.data.get('force', False))  # bypass validation

        if not force:
            # ── Validation sémantique avant toute sauvegarde ──────────────────
            from .alignment_validator import (
                validate_odd_choices, validate_snd30_choices, validate_pcd_prd_choices,
            )

            validation_errors = {}

            if sdg_indicator_codes is not None and len(sdg_indicator_codes) > 0:
                result = validate_odd_choices(project, sdg_indicator_codes)
                if not result['valid']:
                    validation_errors['odd'] = result

            if snd30_axes_numbers is not None and len(snd30_axes_numbers) > 0:
                result = validate_snd30_choices(project, snd30_axes_numbers)
                if not result['valid']:
                    validation_errors['snd30'] = result

            if local_priorities is not None and len(local_priorities) > 0:
                result = validate_pcd_prd_choices(project, local_priorities)
                if not result['valid']:
                    validation_errors['pcd_prd'] = result

            if validation_errors:
                return Response(
                    {'validation_failed': True, **validation_errors},
                    status=status.HTTP_422_UNPROCESSABLE_ENTITY,
                )

        # ── Sauvegarde (validation passée) ────────────────────────────────
        if sdg_indicator_codes is not None:
            ProjectSDGAlignment.objects.filter(project=project).delete()
            for code in sdg_indicator_codes:
                try:
                    indicator = SDGIndicator.objects.get(code=code)
                    ProjectSDGAlignment.objects.create(project=project, indicator=indicator)
                except SDGIndicator.DoesNotExist:
                    continue

        if snd30_axes_numbers is not None:
            ProjectSND30Alignment.objects.filter(project=project).delete()
            for num in snd30_axes_numbers:
                try:
                    axis = SND30Axis.objects.get(number=num)
                    ProjectSND30Alignment.objects.create(project=project, axis=axis)
                except SND30Axis.DoesNotExist:
                    continue

        if local_priorities is not None:
            project.local_priorities = local_priorities
            project.save(update_fields=['local_priorities'])

        project.refresh_from_db()
        project.update_scores(save=True)
        if project.status not in ('draft', 'rejected', 'needs_completion'):
            # Si on modifie un projet déjà soumis/validé, il repasse en brouillon pour ré-audit
            old_status = project.status
            if old_status != 'draft':
                project.status = 'draft'
                project.rejection_reason = ''
                project.save(update_fields=['status', 'rejection_reason'])
                ProjectWorkflowTrace.objects.create(
                    project=project,
                    actor=request.user,
                    from_status=old_status,
                    to_status='draft',
                    comment="Alignement modifié. Le projet doit être à nouveau soumis et validé.",
                )

        return Response({'status': 'alignments synced'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def auto_align(self, request, pk=None):
        roles = user_roles(request.user)
        if roles and not (roles & (ADMIN_ROLES | {ROLE_POINT_FOCAL})):
            return Response(
                {'detail': "Seul un Point Focal ODD ou Administrateur peut effectuer l'alignement automatique."},
                status=status.HTTP_403_FORBIDDEN,
            )

        project = self.get_object()

        from .alignment_validator import (
            _suggest_odd, _tokenize, _project_text, _normalize,
            _SND30_AXES, _PCD_PRD_PRIORITIES, suggest_pcd_objectives,
            RELEVANCE_THRESHOLD, MIN_INDICATORS,
        )

        proj_tokens = _tokenize(_project_text(project))
        proj_norm = _normalize(_project_text(project))

        # 1. MAPPING ODD — ne retient que les suggestions réellement pertinentes
        # (même seuil que la validation manuelle) ; si trop peu franchissent le
        # seuil, on complète avec les meilleures disponibles pour respecter le
        # minimum requis par la validation de complétude.
        suggested_odds = _suggest_odd(proj_tokens, [], project=project)
        odd_codes = [s['code'] for s in suggested_odds if s['score_pct'] >= RELEVANCE_THRESHOLD * 100]
        if len(odd_codes) < MIN_INDICATORS:
            odd_codes = [s['code'] for s in suggested_odds[:MIN_INDICATORS]]

        # 2. MAPPING SND30 — uniquement les axes ayant au moins un mot-clé en
        # commun avec le projet ; pas de sélection forcée sans lien thématique.
        candidates_snd30 = []
        for num, axis in _SND30_AXES.items():
            matches = [kw for kw in axis['keywords'] if kw in proj_norm]
            candidates_snd30.append((len(matches), num))
        candidates_snd30.sort(reverse=True)
        snd30_nums = [num for count, num in candidates_snd30[:2] if count > 0]

        # 3. MAPPING PCD/PRD — utilise les vraies données PCD si disponibles,
        # sinon les catégories génériques — jamais de sélection sans lien.
        pcd_suggestions = suggest_pcd_objectives(project)
        if pcd_suggestions:
            pcd_ids = [s['priority_id'] for s in pcd_suggestions[:2] if s['score_pct'] >= RELEVANCE_THRESHOLD * 100]
        else:
            candidates_pcd = []
            for pid, priority in _PCD_PRD_PRIORITIES.items():
                matches = [kw for kw in priority['keywords'] if kw in proj_norm]
                candidates_pcd.append((len(matches), pid))
            candidates_pcd.sort(reverse=True)
            pcd_ids = [pid for count, pid in candidates_pcd[:2] if count > 0]

        # Persistence
        if odd_codes:
            ProjectSDGAlignment.objects.filter(project=project).delete()
            for code in odd_codes:
                try:
                    indicator = SDGIndicator.objects.get(code=code)
                    ProjectSDGAlignment.objects.create(project=project, indicator=indicator)
                except SDGIndicator.DoesNotExist:
                    continue

        if snd30_nums:
            ProjectSND30Alignment.objects.filter(project=project).delete()
            for num in snd30_nums:
                try:
                    axis = SND30Axis.objects.get(number=num)
                    ProjectSND30Alignment.objects.create(project=project, axis=axis)
                except SND30Axis.DoesNotExist:
                    continue

        if pcd_ids:
            project.local_priorities = pcd_ids
            project.save(update_fields=['local_priorities'])

        project.refresh_from_db()
        project.update_scores(save=True)
        # Les alignements viennent d'être réécrits : quel que soit l'état du
        # dossier (en circuit de validation ou déjà certifié), il doit repasser
        # par une nouvelle soumission plutôt que de garder un statut obsolète
        # sur des alignements qui ont changé sous ses pieds.
        old_status = project.status
        if old_status != 'draft':
            project.status = 'draft'
            project.rejection_reason = ''
            project.is_published = False
            project.published_at = None
            project.save(update_fields=['status', 'rejection_reason', 'is_published', 'published_at'])
            ProjectWorkflowTrace.objects.create(
                project=project,
                actor=request.user,
                from_status=old_status,
                to_status='draft',
                comment="Alignement automatique effectué. Le projet doit être à nouveau soumis et validé.",
            )

        return Response({'status': 'auto_aligned'}, status=status.HTTP_200_OK)
    @action(detail=True, methods=['get'])
    def history(self, request, pk=None):
        """Retourne le journal des transitions de statut du projet."""
        project = self.get_object()
        traces = project.workflow_history.all()
        return Response(WorkflowTraceSerializer(traces, many=True).data)

    @action(detail=True, methods=['get', 'post'], url_path='gadd-evaluation')
    def gadd_evaluation(self, request, pk=None):
        """
        GET  — retourne la fiche d'évaluation GADD du projet (ou null si absente).
        POST — deux usages indépendants, combinables dans le même appel :
          - Portée : { "theme_ids": [<id>, ...] } — thèmes jugés pertinents pour CE
            projet (chaque projet a ses propres objectifs à évaluer, pas les 166).
          - Réponses : { "answers": [ { "objective": <id>, "importance": 1-3|null,
                                         "evaluation_pct": 0-100|null,
                                         "justification": "...", "actions": "..." }, ... ] }
            En pratique le frontend soumet les réponses d'une seule dimension à la
            fois (une page/onglet par dimension).
        """
        project = self.get_object()

        if request.method == 'GET':
            evaluation = getattr(project, 'gadd_evaluation', None)
            if not evaluation:
                # Pas de corps JSON valide pour `None` (DRF renvoie des octets vides) —
                # 204 signale explicitement "pas encore de fiche" au frontend.
                return Response(status=status.HTTP_204_NO_CONTENT)
            return Response(GaddEvaluationSerializer(evaluation).data)

        roles = user_roles(request.user)
        allowed = ADMIN_ROLES | {ROLE_POINT_FOCAL, ROLE_RESPONSABLE_PROJET}
        if roles and not (roles & allowed):
            return Response(
                {'detail': "Seul un Responsable Projet, Point Focal ODD ou Administrateur peut évaluer ce projet."},
                status=status.HTTP_403_FORBIDDEN,
            )

        theme_ids    = request.data.get('theme_ids')
        answers_data = request.data.get('answers')

        if theme_ids is None and not answers_data:
            return Response(
                {'detail': "Fournissez 'theme_ids' et/ou 'answers'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        evaluation, _ = GaddEvaluation.objects.get_or_create(project=project)
        evaluation.evaluated_by = request.user

        update_fields = ['evaluated_by', 'evaluated_at']
        if theme_ids is not None:
            if not isinstance(theme_ids, list):
                return Response({'detail': "'theme_ids' doit être une liste."}, status=status.HTTP_400_BAD_REQUEST)
            evaluation.relevant_theme_ids = theme_ids
            update_fields.append('relevant_theme_ids')
        evaluation.save(update_fields=update_fields)

        for item in (answers_data or []):
            objective_id = item.get('objective')
            if not objective_id:
                continue
            try:
                objective = GaddObjective.objects.get(pk=objective_id)
            except GaddObjective.DoesNotExist:
                continue

            GaddObjectiveAnswer.objects.update_or_create(
                evaluation=evaluation,
                objective=objective,
                defaults={
                    'importance':      item.get('importance'),
                    'evaluation_pct':  item.get('evaluation_pct'),
                    'justification':   item.get('justification') or '',
                    'actions':         item.get('actions') or '',
                },
            )

        evaluation.refresh_from_db()
        return Response(GaddEvaluationSerializer(evaluation).data)

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """
        Import de projets depuis un fichier .xlsx ou .csv.
        Colonnes attendues (en-têtes ligne 1) :
          nom* | description | territoire | budget | date_debut | date_fin | secteur | beneficiaires
        """
        from rest_framework.parsers import MultiPartParser
        import csv, io, decimal

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Aucun fichier fourni.'}, status=status.HTTP_400_BAD_REQUEST)

        roles = user_roles(request.user)
        allowed = ADMIN_ROLES | {ROLE_POINT_FOCAL, ROLE_RESPONSABLE_PROJET}
        if roles and not (roles & allowed):
            return Response({'detail': 'Accès non autorisé.'}, status=status.HTTP_403_FORBIDDEN)

        filename = file.name.lower()
        rows = []

        try:
            if filename.endswith('.csv'):
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)

            elif filename.endswith(('.xlsx', '.xls')):
                import openpyxl
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(cell.value or '').strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                return Response(
                    {'detail': 'Format non supporté. Utilisez .xlsx ou .csv'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            return Response({'detail': f'Erreur lecture fichier : {e}'}, status=status.HTTP_400_BAD_REQUEST)

        created, errors = [], []

        for i, row in enumerate(rows, start=2):
            line = {k.strip().lower(): (str(v).strip() if v is not None else '') for k, v in row.items()}
            nom = line.get('nom') or line.get('name') or ''
            if not nom:
                errors.append(f'Ligne {i} ignorée — colonne « nom » manquante.')
                continue

            try:
                sector_obj = None
                sector_name = line.get('secteur') or line.get('sector') or ''
                if sector_name:
                    from .models import Sector as SectorModel
                    sector_obj = SectorModel.objects.filter(name__iexact=sector_name).first()

                budget_raw = line.get('budget', '') or '0'
                try:
                    budget = decimal.Decimal(str(budget_raw).replace(' ', '').replace(',', '.'))
                except Exception:
                    budget = decimal.Decimal('0')

                commune_obj = None
                territoire = line.get('territoire') or line.get('territory') or ''
                if territoire:
                    from locations.models import Commune
                    commune_obj = Commune.objects.filter(name__iexact=territoire).first()

                def parse_date(val):
                    if not val: return None
                    from django.utils.dateparse import parse_date as _pd
                    return _pd(str(val)[:10])

                beneficiaires_raw = line.get('beneficiaires') or line.get('beneficiaries') or ''
                try:
                    beneficiaires = int(float(beneficiaires_raw)) if beneficiaires_raw else None
                except Exception:
                    beneficiaires = None

                project = Project.objects.create(
                    name=nom,
                    description=line.get('description') or '',
                    territory=territoire,
                    budget=budget,
                    start_date=parse_date(line.get('date_debut') or line.get('start_date')),
                    end_date=parse_date(line.get('date_fin') or line.get('end_date')),
                    sector=sector_obj,
                    commune=commune_obj,
                    beneficiaries_count=beneficiaires,
                    status='draft',
                    owner=request.user,
                )
                created.append({'id': project.id, 'name': project.name})

            except Exception as e:
                errors.append(f'Ligne {i} ({nom}) — erreur : {e}')

        return Response({
            'created': len(created),
            'projects': created,
            'errors': errors,
            'message': f'{len(created)} projet(s) importé(s). {len(errors)} erreur(s).',
        })

    @action(detail=False, methods=['get'], url_path='template-excel')
    def template_excel(self, request):
        """Téléchargement du modèle Excel pour l'import."""
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Projets'

        headers = ['nom', 'description', 'territoire', 'budget', 'date_debut', 'date_fin', 'secteur', 'beneficiaires']
        ws.append(headers)

        # Ligne exemple
        ws.append([
            'Adduction eau potable Bafoussam',
            'Construction de 3 forages dans les quartiers défavorisés',
            'Bafoussam II',
            '45000000',
            '2026-03-01',
            '2026-12-31',
            'Eau et assainissement',
            '5000',
        ])

        # Style en-têtes
        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1A237E')

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 50
        for col in 'CDEFGH':
            ws.column_dimensions[col].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="modele_import_projets.xlsx"'
        wb.save(response)
        return response


class PCDObjectiveViewSet(viewsets.ReadOnlyModelViewSet):
    """
    GET /api/odd/pcd-objectives/                         — tous les objectifs PCD
    GET /api/odd/pcd-objectives/?commune=Bafoussam I     — filtrés par commune
    GET /api/odd/pcd-objectives/?project=<id>            — filtrés par commune du projet
    """
    serializer_class   = PCDObjectiveSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = PCDObjective.objects.all()
        commune_nom = self.request.query_params.get('commune')
        project_id  = self.request.query_params.get('project')

        if project_id:
            try:
                project = Project.objects.get(pk=project_id)
                nom = (project.commune.name if project.commune else None) or project.territory
                if nom:
                    qs = qs.filter(commune_nom__icontains=nom.split()[0])
            except Project.DoesNotExist:
                pass
        elif commune_nom:
            qs = qs.filter(commune_nom__icontains=commune_nom.split()[0])

        return qs


class PreuveViewSet(viewsets.ModelViewSet):
    """
    GET    /api/odd/preuves/?project=<id>  — liste les preuves d'un projet
    POST   /api/odd/preuves/               — dépôt d'une preuve (multipart)
    DELETE /api/odd/preuves/<id>/          — suppression
    """
    serializer_class   = PreuveSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class   = None

    def get_queryset(self):
        qs = Preuve.objects.select_related('uploaded_by', 'project').all()
        project_id = self.request.query_params.get('project')
        if project_id:
            qs = qs.filter(project_id=project_id)
        return qs

    def perform_create(self, serializer):
        fichier = self.request.FILES.get('fichier')
        hash_sha256 = ''
        taille = 0
        if fichier:
            sha256 = hashlib.sha256()
            for chunk in fichier.chunks():
                sha256.update(chunk)
            hash_sha256 = sha256.hexdigest()
            taille = fichier.size
            fichier.seek(0)

        serializer.save(
            uploaded_by=self.request.user,
            hash_sha256=hash_sha256,
            taille=taille,
        )

    def destroy(self, request, *args, **kwargs):
        preuve = self.get_object()
        # Seul le déposant ou un admin peut supprimer
        roles = user_roles(request.user)
        if preuve.uploaded_by != request.user and not (roles & ADMIN_ROLES or request.user.is_staff):
            return Response(
                {'detail': 'Vous ne pouvez supprimer que vos propres preuves.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        # Supprimer le fichier physique
        if preuve.fichier:
            preuve.fichier.delete(save=False)
        return super().destroy(request, *args, **kwargs)


class SDGViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SDG.objects.all().order_by('number')
    serializer_class = SDGSerializer
    permission_classes = [permissions.AllowAny]
    pagination_class = None


class SND30AxisViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = SND30Axis.objects.all().order_by('number')
    serializer_class = SND30AxisSerializer
    permission_classes = [permissions.AllowAny]
    pagination_class = None


class GaddDimensionViewSet(viewsets.ReadOnlyModelViewSet):
    """Référentiel GADD (6 dimensions, thèmes, 166 objectifs officiels) — lecture seule."""
    queryset = GaddDimension.objects.all().prefetch_related('themes__objectives').order_by('order')
    serializer_class = GaddDimensionSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None


# ──────────────────────────────────────────────────────────────────────────
# Rôles ODD — demande + validation
# ──────────────────────────────────────────────────────────────────────────

class MyOddRoleView(APIView):
    """GET /api/odd/my-role/ — retourne les rôles ODD actifs + la demande en cours."""
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        roles = list(user_roles(request.user))
        try:
            req = OddRoleRequest.objects.get(user=request.user)
            request_data = OddRoleRequestSerializer(req).data
        except OddRoleRequest.DoesNotExist:
            request_data = None

        return Response({'roles': roles, 'request': request_data})


class OddRoleRequestViewSet(viewsets.ModelViewSet):
    """
    POST   /api/odd/role-requests/           — l'utilisateur soumet sa demande
    GET    /api/odd/role-requests/            — admin : liste toutes les demandes
    GET    /api/odd/role-requests/{id}/       — détail
    POST   /api/odd/role-requests/{id}/approve/ — admin : approuve
    POST   /api/odd/role-requests/{id}/reject/  — admin : rejette
    """
    serializer_class = OddRoleRequestSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if not getattr(user, 'is_authenticated', False):
            return OddRoleRequest.objects.none()

        roles = user_roles(user)
        if roles & ADMIN_ROLES or getattr(user, 'is_staff', False):
            return OddRoleRequest.objects.select_related('user', 'reviewed_by').all()
        return OddRoleRequest.objects.filter(user=user)

    def create(self, request, *args, **kwargs):
        # Un utilisateur ne peut avoir qu'une seule demande active
        if OddRoleRequest.objects.filter(user=request.user).exists():
            existing = OddRoleRequest.objects.get(user=request.user)
            if existing.status == 'pending':
                return Response(
                    {'detail': 'Vous avez déjà une demande en attente.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # Si rejetée ou approuvée, on met à jour
            existing.requested_role = request.data.get('requested_role', existing.requested_role)
            existing.motivation     = request.data.get('motivation', existing.motivation)
            existing.status         = 'pending'
            existing.reviewed_at    = None
            existing.reviewed_by    = None
            existing.review_comment = ''
            existing.save()
            return Response(OddRoleRequestSerializer(existing).data, status=status.HTTP_200_OK)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(user=request.user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Admin approuve la demande → assigne le rôle en base."""
        if not (user_roles(request.user) & ADMIN_ROLES or request.user.is_staff):
            return Response({'detail': 'Action réservée aux administrateurs.'}, status=status.HTTP_403_FORBIDDEN)

        role_request = self.get_object()
        if role_request.status == 'approved':
            return Response({'detail': 'Cette demande est déjà approuvée.'}, status=status.HTTP_400_BAD_REQUEST)

        # Assigner le rôle dans accounts.Role
        from accounts.models import Role as AccountRole
        role_obj, _ = AccountRole.objects.get_or_create(name=role_request.requested_role)
        role_request.user.roles.add(role_obj)

        role_request.status      = 'approved'
        role_request.reviewed_at = timezone.now()
        role_request.reviewed_by = request.user
        role_request.review_comment = request.data.get('comment', '')
        role_request.save()

        return Response({'detail': f"Rôle '{role_request.requested_role}' assigné à {role_request.user.email}."})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):  # noqa (defined below)
        """Admin rejette la demande."""
        if not (user_roles(request.user) & ADMIN_ROLES or request.user.is_staff):
            return Response({'detail': 'Action réservée aux administrateurs.'}, status=status.HTTP_403_FORBIDDEN)

        role_request = self.get_object()
        if role_request.status == 'rejected':
            return Response({'detail': 'Cette demande est déjà rejetée.'}, status=status.HTTP_400_BAD_REQUEST)

        comment = request.data.get('comment', '').strip()
        if not comment:
            return Response({'detail': 'Un motif de rejet est obligatoire.'}, status=status.HTTP_400_BAD_REQUEST)

        role_request.status         = 'rejected'
        role_request.reviewed_at    = timezone.now()
        role_request.reviewed_by    = request.user
        role_request.review_comment = comment
        role_request.save()

        return Response({'detail': f"Demande de {role_request.user.email} rejetée."})


# ──────────────────────────────────────────────────────────────────────────
# Périodes de collecte
# ──────────────────────────────────────────────────────────────────────────

class PeriodViewSet(viewsets.ModelViewSet):
    queryset = Period.objects.all()
    serializer_class = PeriodSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    def destroy(self, request, *args, **kwargs):
        period = self.get_object()
        if not is_admin(request.user):
            return Response({'detail': 'Seul un administrateur peut supprimer une période.'},
                            status=status.HTTP_403_FORBIDDEN)
        if period.mesures.exists():
            return Response({'detail': 'Cette période contient des mesures — suppression impossible.'},
                            status=status.HTTP_400_BAD_REQUEST)
        return super().destroy(request, *args, **kwargs)


# ──────────────────────────────────────────────────────────────────────────
# Mesures KPI (collecte périodique)
# ──────────────────────────────────────────────────────────────────────────

class MesureKPIViewSet(viewsets.ModelViewSet):
    serializer_class = MesureKPISerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        if getattr(self, 'swagger_fake_view', False):
            return MesureKPI.objects.none()

        user = self.request.user
        qs = MesureKPI.objects.select_related(
            'indicator__target__sdg', 'period', 'saisi_par', 'project'
        ).all()

        project_id = self.request.query_params.get('project')
        period_id  = self.request.query_params.get('period')
        sdg_number = self.request.query_params.get('sdg')
        statut     = self.request.query_params.get('statut')

        if project_id:
            qs = qs.filter(project_id=project_id)
        if period_id:
            qs = qs.filter(period_id=period_id)
        if sdg_number:
            qs = qs.filter(indicator__target__sdg__number=sdg_number)
        if statut:
            qs = qs.filter(statut=statut)

        if not ROLES_ENABLED:
            return qs

        roles = user_roles(user)
        if roles and not (roles & ADMIN_ROLES or roles & VALIDATOR_ROLES):
            qs = qs.filter(project__owner=user)

        return qs

    def perform_create(self, serializer):
        serializer.save(saisi_par=self.request.user)

    def perform_update(self, serializer):
        serializer.save(saisi_par=self.request.user)

    def create(self, request, *args, **kwargs):
        roles = user_roles(request.user)
        allowed = ADMIN_ROLES | {ROLE_POINT_FOCAL, ROLE_RESPONSABLE_PROJET}
        if roles and not (roles & allowed):
            return Response(
                {'detail': 'Seul un Responsable Projet, Point Focal ou Administrateur peut saisir des mesures.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().create(request, *args, **kwargs)

    @action(detail=False, methods=['get'], url_path='template-excel')
    def template_excel(self, request):
        """Téléchargement du modèle Excel pour la collecte des KPI."""
        import openpyxl
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Saisie_KPI'

        headers = ['projet_id', 'indicateur_code', 'periode_id', 'valeur', 'valeur_cible', 'commentaire']
        ws.append(headers)

        ws.append([
            1,
            '3.1.1',
            1,
            '450',
            '500',
            'Saisie trimestrielle',
        ])

        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1A237E')

        for col in 'ABCDEF':
            ws.column_dimensions[col].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="modele_import_kpi.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """Import en masse de mesures KPI via Excel/CSV."""
        import csv, io, decimal

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'Aucun fichier fourni.'}, status=status.HTTP_400_BAD_REQUEST)

        roles = user_roles(request.user)
        allowed = ADMIN_ROLES | {ROLE_POINT_FOCAL, ROLE_RESPONSABLE_PROJET}
        if roles and not (roles & allowed):
            return Response({'detail': 'Accès non autorisé.'}, status=status.HTTP_403_FORBIDDEN)

        filename = file.name.lower()
        rows = []

        try:
            if filename.endswith('.csv'):
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif filename.endswith(('.xlsx', '.xls')):
                import openpyxl
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(cell.value or '').strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                return Response(
                    {'detail': 'Format non supporté. Utilisez .xlsx ou .csv'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except Exception as e:
            return Response({'detail': f'Erreur lecture fichier : {e}'}, status=status.HTTP_400_BAD_REQUEST)

        created, errors = 0, []

        for i, row in enumerate(rows, start=2):
            line = {k.strip().lower(): (str(v).strip() if v is not None else '') for k, v in row.items()}
            
            project_raw = line.get('projet_id') or ''
            indicator_raw = line.get('indicateur_code') or ''
            period_raw = line.get('periode_id') or ''
            
            if not project_raw or not indicator_raw or not period_raw:
                errors.append(f'Ligne {i} ignorée — identifiants manquants (projet_id, indicateur_code, periode_id).')
                continue

            try:
                project = Project.objects.get(pk=int(float(project_raw)))
                indicator = SDGIndicator.objects.get(code=indicator_raw)
                period = Period.objects.get(pk=int(float(period_raw)))
                
                if not is_admin(request.user) and project.owner != request.user:
                    errors.append(f'Ligne {i} — vous ne pouvez pas modifier les données du projet {project.id}.')
                    continue

                valeur_raw = line.get('valeur') or ''
                valeur_cible_raw = line.get('valeur_cible') or ''
                
                try:
                    valeur = decimal.Decimal(valeur_raw.replace(' ', '').replace(',', '.')) if valeur_raw else None
                except Exception:
                    valeur = None
                    
                try:
                    valeur_cible = decimal.Decimal(valeur_cible_raw.replace(' ', '').replace(',', '.')) if valeur_cible_raw else None
                except Exception:
                    valeur_cible = None

                commentaire = line.get('commentaire') or ''

                mesure, is_new = MesureKPI.objects.update_or_create(
                    project=project,
                    indicator=indicator,
                    period=period,
                    defaults={
                        'valeur': valeur,
                        'valeur_cible': valeur_cible,
                        'commentaire': commentaire,
                        'saisi_par': request.user,
                        'statut': 'draft',
                    }
                )
                created += 1

            except Project.DoesNotExist:
                errors.append(f'Ligne {i} — Projet {project_raw} introuvable.')
            except SDGIndicator.DoesNotExist:
                errors.append(f'Ligne {i} — Indicateur {indicator_raw} introuvable.')
            except Period.DoesNotExist:
                errors.append(f'Ligne {i} — Période {period_raw} introuvable.')
            except Exception as e:
                errors.append(f'Ligne {i} — erreur : {e}')

        return Response({
            'created': created,
            'errors': errors,
            'message': f'{created} mesure(s) importée(s)/mise(s) à jour. {len(errors)} erreur(s).',
        })

    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Soumet une mesure pour validation."""
        mesure = self.get_object()
        if mesure.statut != 'draft':
            return Response({'detail': 'Seules les mesures en brouillon peuvent être soumises.'},
                            status=status.HTTP_400_BAD_REQUEST)
        mesure.statut = 'submitted'
        mesure.save(update_fields=['statut'])
        return Response({'detail': 'Mesure soumise pour validation.'})

    @action(detail=True, methods=['post'])
    def validate(self, request, pk=None):
        """Valide une mesure (validateur/admin seulement)."""
        if ROLES_ENABLED:
            roles = user_roles(request.user)
            if not (roles & ADMIN_ROLES or roles & VALIDATOR_ROLES):
                return Response({'detail': 'Action réservée aux validateurs.'},
                                status=status.HTTP_403_FORBIDDEN)
        mesure = self.get_object()
        mesure.statut = 'validated'
        mesure.save(update_fields=['statut'])
        return Response({'detail': 'Mesure validée.'})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Rejette une mesure avec motif."""
        if ROLES_ENABLED:
            roles = user_roles(request.user)
            if not (roles & ADMIN_ROLES or roles & VALIDATOR_ROLES):
                return Response({'detail': 'Action réservée aux validateurs.'},
                                status=status.HTTP_403_FORBIDDEN)
        comment = request.data.get('comment', '').strip()
        if not comment:
            return Response({'detail': 'Un motif de rejet est obligatoire.'},
                            status=status.HTTP_400_BAD_REQUEST)
        mesure = self.get_object()
        mesure.statut = 'rejected'
        mesure.commentaire = comment
        mesure.save(update_fields=['statut', 'commentaire'])
        return Response({'detail': 'Mesure rejetée.'})


# ──────────────────────────────────────────────────────────────────────────
# Statistiques par ODD et par Secteur
# ──────────────────────────────────────────────────────────────────────────

class StatsODDView(APIView):
    """
    GET /api/odd/stats/by-odd/
    Retourne pour chaque ODD : nombre de projets, budget total, bénéficiaires.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from django.db.models import Count, Sum
        from .models import ProjectSDGAlignment

        user  = request.user
        roles = user_roles(user)

        if is_admin(user) or not roles:
            projects_qs = Project.objects.all()
        elif roles & VALIDATOR_ROLES:
            projects_qs = Project.objects.filter(
                status__in=['submitted', 'validated_ctd', 'controlling_arrdel', 'validated', 'rejected']
            )
        else:
            projects_qs = Project.objects.filter(owner=user)

        sdgs = SDG.objects.all().prefetch_related('targets__indicators').order_by('number')
        result = []

        for sdg in sdgs:
            indicator_ids = list(
                SDGIndicator.objects.filter(target__sdg=sdg).values_list('id', flat=True)
            )
            aligned_project_ids = list(
                ProjectSDGAlignment.objects
                .filter(indicator_id__in=indicator_ids, project__in=projects_qs)
                .values_list('project_id', flat=True)
                .distinct()
            )
            aligned_projects = projects_qs.filter(id__in=aligned_project_ids)
            agg = aligned_projects.aggregate(
                total_budget=Sum('budget'),
                total_beneficiaries=Sum('beneficiaries_count'),
            )
            status_breakdown = {}
            for p in aligned_projects.values('status'):
                s = p['status']
                status_breakdown[s] = status_breakdown.get(s, 0) + 1

            result.append({
                'sdg_number':           sdg.number,
                'sdg_name':             sdg.name,
                'sdg_color':            sdg.color,
                'project_count':        len(aligned_project_ids),
                'validated_count':      status_breakdown.get('validated', 0),
                'total_budget':         float(agg['total_budget'] or 0),
                'total_beneficiaries':  agg['total_beneficiaries'] or 0,
                'status_breakdown':     status_breakdown,
            })

        return Response(result)


class StatsSecteurView(APIView):
    """
    GET /api/odd/stats/by-sector/
    Retourne pour chaque secteur : nombre de projets, budget, statuts.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        from django.db.models import Sum

        user  = request.user
        roles = user_roles(user)

        if is_admin(user) or not roles:
            projects_qs = Project.objects.select_related('sector').all()
        elif roles & VALIDATOR_ROLES:
            projects_qs = Project.objects.select_related('sector').filter(
                status__in=['submitted', 'validated_ctd', 'controlling_arrdel', 'validated', 'rejected']
            )
        else:
            projects_qs = Project.objects.select_related('sector').filter(owner=user)

        sectors = Sector.objects.all().order_by('name')
        result  = []

        for sec in sectors:
            sec_projects = projects_qs.filter(sector=sec)
            agg = sec_projects.aggregate(
                total_budget=Sum('budget'),
                total_beneficiaries=Sum('beneficiaries_count'),
            )
            status_breakdown = {}
            for p in sec_projects.values('status'):
                s = p['status']
                status_breakdown[s] = status_breakdown.get(s, 0) + 1

            if sec_projects.exists():
                result.append({
                    'sector_id':            sec.id,
                    'sector_name':          sec.name,
                    'project_count':        sec_projects.count(),
                    'validated_count':      status_breakdown.get('validated', 0),
                    'total_budget':         float(agg['total_budget'] or 0),
                    'total_beneficiaries':  agg['total_beneficiaries'] or 0,
                    'status_breakdown':     status_breakdown,
                })

        no_sector = projects_qs.filter(sector__isnull=True)
        if no_sector.exists():
            agg = no_sector.aggregate(
                total_budget=Sum('budget'),
                total_beneficiaries=Sum('beneficiaries_count'),
            )
            status_breakdown = {}
            for p in no_sector.values('status'):
                s = p['status']
                status_breakdown[s] = status_breakdown.get(s, 0) + 1
            result.append({
                'sector_id':   None,
                'sector_name': 'Non classifié',
                'project_count': no_sector.count(),
                'validated_count': status_breakdown.get('validated', 0),
                'total_budget': float(agg['total_budget'] or 0),
                'total_beneficiaries': agg['total_beneficiaries'] or 0,
                'status_breakdown': status_breakdown,
            })

        return Response(result)


class StatsCommuneView(APIView):
    """
    GET /api/odd/stats/by-commune/
    Retourne pour chaque commune : nombre de projets, % d'alignement global moyen, budget, statuts.
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user  = request.user
        roles = user_roles(user)

        base_qs = Project.objects.select_related('commune').prefetch_related(
            'sdg_alignments__indicator__target__sdg', 'snd30_alignments'
        )
        if is_admin(user) or not roles:
            projects_qs = base_qs.all()
        elif roles & VALIDATOR_ROLES:
            projects_qs = base_qs.filter(
                status__in=['submitted', 'validated_ctd', 'controlling_arrdel', 'validated', 'rejected']
            )
        else:
            projects_qs = base_qs.filter(owner=user)

        # Regroupement par commune liée (FK) si disponible, sinon par territoire libre
        groups = {}
        for p in projects_qs:
            # Recalcule le score d'alignement en mémoire (sans persister) pour que la
            # moyenne reflète l'état réel des alignements, y compris pour les projets
            # jamais validés/exportés (dont le score_global stocké serait figé à 0).
            p.update_scores(save=False)
            if p.commune_id:
                key   = f"commune:{p.commune_id}"
                label = p.commune.name
            else:
                territory = (p.territory or '').strip()
                if not territory:
                    continue
                key   = f"territory:{territory.lower()}"
                label = territory
            g = groups.setdefault(key, {'commune_id': p.commune_id, 'label': label, 'projects': []})
            g['projects'].append(p)

        result = []
        for g in groups.values():
            projs = g['projects']
            status_breakdown = {}
            for p in projs:
                status_breakdown[p.status] = status_breakdown.get(p.status, 0) + 1

            total_budget        = sum(float(p.budget or 0) for p in projs)
            total_beneficiaries = sum(p.beneficiaries_count or 0 for p in projs)
            scored_values        = [float(p.score_global or 0) for p in projs if p.score_global]
            avg_alignment_score  = round(sum(scored_values) / len(scored_values), 1) if scored_values else 0

            result.append({
                'commune_id':          g['commune_id'],
                'commune_name':        g['label'],
                'project_count':       len(projs),
                'validated_count':     status_breakdown.get('validated', 0),
                'avg_alignment_score': avg_alignment_score,
                'total_budget':        total_budget,
                'total_beneficiaries': total_beneficiaries,
                'status_breakdown':    status_breakdown,
            })

        result.sort(key=lambda r: r['commune_name'])
        return Response(result)


class CommuneProjectsView(APIView):
    """
    GET /api/odd/stats/by-commune/projects/?commune_id=5
    GET /api/odd/stats/by-commune/projects/?territory=Bafoussam

    Retourne, pour une commune donnée, la liste de ses projets avec, pour
    chacun, le détail du pourcentage d'alignement par ODD (nombre d'indicateurs
    couverts par le projet / nombre total d'indicateurs du référentiel pour cet ODD).
    """
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        user  = request.user
        roles = user_roles(user)

        base_qs = Project.objects.select_related('commune', 'sector').prefetch_related(
            'sdg_alignments__indicator__target__sdg'
        )
        if is_admin(user) or not roles:
            projects_qs = base_qs.all()
        elif roles & VALIDATOR_ROLES:
            projects_qs = base_qs.filter(
                status__in=['submitted', 'validated_ctd', 'controlling_arrdel', 'validated', 'rejected']
            )
        else:
            projects_qs = base_qs.filter(owner=user)

        commune_id = request.query_params.get('commune_id')
        territory  = request.query_params.get('territory')

        if commune_id:
            projects_qs = projects_qs.filter(commune_id=commune_id)
        elif territory:
            projects_qs = projects_qs.filter(commune__isnull=True, territory__iexact=territory.strip())
        else:
            return Response({'detail': 'Paramètre commune_id ou territory requis.'}, status=status.HTTP_400_BAD_REQUEST)

        # Référentiel : nombre total d'indicateurs par ODD, pour normaliser les pourcentages
        sdg_meta = {}
        for sdg in SDG.objects.all():
            sdg_meta[sdg.number] = {
                'name':             sdg.name,
                'color':            sdg.color,
                'total_indicators': SDGIndicator.objects.filter(target__sdg=sdg).count(),
            }

        result = []
        for p in projects_qs:
            p.update_scores(save=False)

            indicators_by_sdg = {}
            for al in p.sdg_alignments.all():
                num = al.indicator.target.sdg.number
                indicators_by_sdg.setdefault(num, set()).add(al.indicator_id)

            odd_breakdown = []
            for num, indicator_ids in sorted(indicators_by_sdg.items()):
                meta  = sdg_meta.get(num, {'name': '', 'color': '#6b7280', 'total_indicators': 0})
                total = meta['total_indicators'] or 1
                pct   = round(min(100, (len(indicator_ids) / total) * 100))
                odd_breakdown.append({
                    'sdg_number':         num,
                    'sdg_name':           meta['name'],
                    'sdg_color':          meta['color'],
                    'aligned_indicators': len(indicator_ids),
                    'total_indicators':   meta['total_indicators'],
                    'percentage':         pct,
                })

            result.append({
                'id':            p.id,
                'name':          p.name,
                'status':        p.status,
                'status_label':  p.get_status_display(),
                'territory':     p.territory,
                'sector_name':   p.sector.name if p.sector else None,
                'budget':        float(p.budget or 0),
                'score_global':  p.score_global,
                'grade':         p.grade,
                'odd_breakdown': odd_breakdown,
            })

        result.sort(key=lambda r: r['name'])
        return Response(result)


# ──────────────────────────────────────────────────────────────────────────
# Campagnes de Reporting
# ──────────────────────────────────────────────────────────────────────────

class CampagneReportingViewSet(viewsets.ModelViewSet):
    serializer_class   = CampagneReportingSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class   = None

    def get_queryset(self):
        return CampagneReporting.objects.all()

    def perform_create(self, serializer):
        if not is_admin(self.request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Seul un administrateur peut créer une campagne.')
        campagne = serializer.save(created_by=self.request.user)

        # Créer automatiquement la Period correspondante si elle n'existe pas déjà
        year    = campagne.year
        quarter = campagne.quarter
        label   = campagne.label
        period, created = Period.objects.get_or_create(
            year=year,
            quarter=quarter,
            defaults={'label': label},
        )
        # Si la period existait déjà mais sans label lisible, on la met à jour
        if not created and not period.label:
            period.label = label
            period.save(update_fields=['label'])

    @action(detail=True, methods=['post'])
    def open(self, request, pk=None):
        if not is_admin(request.user):
            return Response({'detail': 'Action réservée aux administrateurs.'}, status=status.HTTP_403_FORBIDDEN)
        campagne = self.get_object()
        campagne.status = 'open'
        campagne.save(update_fields=['status'])
        return Response({'detail': 'Campagne ouverte.'})

    @action(detail=True, methods=['post'])
    def close(self, request, pk=None):
        if not is_admin(request.user):
            return Response({'detail': 'Action réservée aux administrateurs.'}, status=status.HTTP_403_FORBIDDEN)
        campagne = self.get_object()
        campagne.status = 'closed'
        campagne.save(update_fields=['status'])
        return Response({'detail': 'Campagne clôturée.'})


# ──────────────────────────────────────────────────────────────────────────
# Rapports archivés
# ──────────────────────────────────────────────────────────────────────────

class RapportViewSet(viewsets.ModelViewSet):
    serializer_class   = RapportSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class   = None

    def get_queryset(self):
        qs = Rapport.objects.select_related('period', 'campagne', 'generated_by').all()
        rapport_type = self.request.query_params.get('type')
        if rapport_type:
            qs = qs.filter(type=rapport_type)
        return qs

    def perform_create(self, serializer):
        serializer.save(generated_by=self.request.user)

    @action(detail=True, methods=['post'])
    def validate(self, request, pk=None):
        if not is_admin(request.user):
            return Response({'detail': 'Action réservée aux administrateurs.'}, status=status.HTTP_403_FORBIDDEN)
        rapport = self.get_object()
        rapport.status = 'validated'
        rapport.save(update_fields=['status'])
        return Response({'detail': 'Rapport validé.'})

    @action(detail=True, methods=['post'])
    def archive(self, request, pk=None):
        if not is_admin(request.user):
            return Response({'detail': 'Action réservée aux administrateurs.'}, status=status.HTTP_403_FORBIDDEN)
        rapport = self.get_object()
        rapport.status = 'archived'
        rapport.save(update_fields=['status'])
        return Response({'detail': 'Rapport archivé.'})


# ──────────────────────────────────────────────────────────────────────────
# Publication / Open Data
# ──────────────────────────────────────────────────────────────────────────

class PublicationViewSet(viewsets.ViewSet):
    """
    GET  /api/odd/publications/          — liste projets publiés (public, sans auth)
    POST /api/odd/publications/{id}/publish/  — marquer publiable (admin)
    POST /api/odd/publications/{id}/unpublish/ — dépublier (admin)
    """

    def get_permissions(self):
        if self.action in ['list', 'stats', 'retrieve', 'download']:
            return [permissions.AllowAny()]
        return [permissions.IsAuthenticated()]

    def list(self, request):
        projects = Project.objects.filter(is_published=True, status='validated').select_related('sector', 'commune')
        from .serializers import ProjectSerializer as PS
        data = PS(projects, many=True, context={'request': request}).data
        return Response(data)

    def _get_published_project(self, pk):
        return Project.objects.filter(
            pk=pk, is_published=True, status='validated'
        ).select_related('sector', 'commune').first()

    def retrieve(self, request, pk=None):
        project = self._get_published_project(pk)
        if not project:
            return Response({'detail': 'Projet introuvable ou non publié.'}, status=status.HTTP_404_NOT_FOUND)
        from .serializers import PublicProjectDetailSerializer
        data = PublicProjectDetailSerializer(project, context={'request': request}).data
        return Response(data)

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        project = self._get_published_project(pk)
        if not project:
            return Response({'detail': 'Projet introuvable ou non publié.'}, status=status.HTTP_404_NOT_FOUND)
        pdf_content = generate_project_alignment_report(project)
        response = HttpResponse(pdf_content, content_type='application/pdf')
        filename = f"Fiche_Projet_{project.id}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['post'])
    def publish(self, request, pk=None):
        from django.utils import timezone
        if not is_admin(request.user):
            return Response({'detail': 'Action réservée aux administrateurs.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            project = Project.objects.get(pk=pk)
        except Project.DoesNotExist:
            return Response({'detail': 'Projet introuvable.'}, status=status.HTTP_404_NOT_FOUND)
        if project.status != 'validated':
            return Response({'detail': 'Seuls les projets validés peuvent être publiés.'}, status=status.HTTP_400_BAD_REQUEST)
        project.is_published = True
        project.published_at = timezone.now()
        project.save(update_fields=['is_published', 'published_at'])
        return Response({'detail': 'Projet publié.'})

    @action(detail=True, methods=['post'])
    def unpublish(self, request, pk=None):
        if not is_admin(request.user):
            return Response({'detail': 'Action réservée aux administrateurs.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            project = Project.objects.get(pk=pk)
        except Project.DoesNotExist:
            return Response({'detail': 'Projet introuvable.'}, status=status.HTTP_404_NOT_FOUND)
        project.is_published = False
        project.published_at = None
        project.save(update_fields=['is_published', 'published_at'])
        return Response({'detail': 'Projet dépublié.'})

    @action(detail=False, methods=['get'])
    def stats(self, request):
        from django.db.models import Sum

        published_projects = Project.objects.filter(is_published=True, status='validated')
        total_projects = published_projects.count()
        
        total_budget = published_projects.aggregate(total=Sum('budget'))['total'] or 0
        total_beneficiaries = published_projects.aggregate(total=Sum('beneficiaries_count'))['total'] or 0
        
        # Calculate SDG stats
        sdg_distribution = {}
        for project in published_projects:
            # On utilise les alignements liés
            nums = project.sdg_alignments.values_list('indicator__target__sdg__number', flat=True).distinct()
            for num in nums:
                sdg_distribution[num] = sdg_distribution.get(num, 0) + 1

        top_sdgs = sorted(
            [{'sdg_number': k, 'count': v} for k, v in sdg_distribution.items()],
            key=lambda x: x['count'], reverse=True
        )[:5]

        return Response({
            'total_projects': total_projects,
            'total_budget': float(total_budget),
            'total_beneficiaries': total_beneficiaries,
            'top_sdgs': top_sdgs,
        })



