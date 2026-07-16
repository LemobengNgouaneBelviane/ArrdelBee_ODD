"""
Commande d'import des Cadres Logiques PCD depuis les fichiers Excel.
Usage : python manage.py import_pcd
        python manage.py import_pcd --reset  (vide la table avant import)
"""
import os
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from odd.models import PCDObjective
from locations.models import Commune


BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..', '..'))

FILES = {
    'Bafoussam I':   'Cadre Logique BAFOUSSAM1 OK.xlsx',
    'Bafoussam II':  'Cadre Logique BAFOUSSAM II.xlsx',
    'Bafoussam III': 'CADRE LOGIQUE BAFOUSSAM III ACTUALISE.xlsx',
}


def _clean(val):
    if val is None:
        return ''
    return str(val).replace('\n', ' ').replace('\xa0', ' ').strip()


def _budget(val):
    if val is None:
        return None
    try:
        return int(float(str(val).replace(' ', '').replace('\xa0', '')))
    except (ValueError, TypeError):
        return None


def _parse_bafoussam_i(ws):
    """
    Colonnes (0-indexed après en-tête) :
    1=collectivite, 2=secteur_id, 3=programme_id, 4=objectifGlobal,
    5=codeObjGlobal, 7=objectifSpecifique, 17=budget, 18=resultats
    """
    seen = set()
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not any(c is not None for c in row):
            continue
        obj_global = _clean(row[4] if len(row) > 4 else None)
        obj_spec   = _clean(row[7] if len(row) > 7 else None)
        if not obj_global:
            continue
        key = (obj_global, obj_spec)
        if key in seen:
            continue
        seen.add(key)
        rows.append({
            'secteur':              f"Secteur {_clean(row[2] if len(row) > 2 else '')}",
            'programme':            f"Programme {_clean(row[3] if len(row) > 3 else '')}",
            'objectif_global':      obj_global,
            'code_objectif_global': _clean(row[5] if len(row) > 5 else ''),
            'objectif_specifique':  obj_spec,
            'resultats':            _clean(row[18] if len(row) > 18 else None),
            'budget':               _budget(row[17] if len(row) > 17 else None),
        })
    return rows


def _parse_bafoussam_ii(ws):
    """
    Colonnes (0-indexed après en-tête) :
    3=idSecteurs(texte), 4=IdChapitres(texte), 5=IdProgrammes(texte),
    6=Objectifs principaux, 7=Projets(obj_spec), 9=Résultats, 10=Budget
    """
    seen = set()
    rows = []
    current_global = ''
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not any(c is not None for c in row):
            continue
        obj_global = _clean(row[6] if len(row) > 6 else None)
        if obj_global:
            current_global = obj_global
        else:
            obj_global = current_global

        obj_spec = _clean(row[7] if len(row) > 7 else None)
        if not obj_global:
            continue
        key = (obj_global, obj_spec)
        if key in seen:
            continue
        seen.add(key)
        rows.append({
            'secteur':              _clean(row[3] if len(row) > 3 else ''),
            'programme':            _clean(row[5] if len(row) > 5 else ''),
            'objectif_global':      obj_global,
            'code_objectif_global': '',
            'objectif_specifique':  obj_spec,
            'resultats':            _clean(row[9] if len(row) > 9 else None),
            'budget':               _budget(row[10] if len(row) > 10 else None),
        })
    return rows


def _parse_bafoussam_iii(ws):
    """
    Colonnes (0-indexed après en-tête) :
    4=secteur_id, 6=programme_id, 7=objectifGlobal, 8=codeObjGlobal,
    9=objectifSpecifique, 19=resultats, 21=budget
    """
    seen = set()
    rows = []
    current_global = ''
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not any(c is not None for c in row):
            continue
        obj_global = _clean(row[7] if len(row) > 7 else None)
        if obj_global:
            current_global = obj_global
        else:
            obj_global = current_global

        obj_spec = _clean(row[9] if len(row) > 9 else None)
        if not obj_global:
            continue
        key = (obj_global, obj_spec)
        if key in seen:
            continue
        seen.add(key)
        rows.append({
            'secteur':              f"Secteur {_clean(row[4] if len(row) > 4 else '')}",
            'programme':            f"Programme {_clean(row[6] if len(row) > 6 else '')}",
            'objectif_global':      obj_global,
            'code_objectif_global': _clean(row[8] if len(row) > 8 else ''),
            'objectif_specifique':  obj_spec,
            'resultats':            _clean(row[19] if len(row) > 19 else None),
            'budget':               _budget(row[21] if len(row) > 21 else None),
        })
    return rows


PARSERS = {
    'Bafoussam I':   _parse_bafoussam_i,
    'Bafoussam II':  _parse_bafoussam_ii,
    'Bafoussam III': _parse_bafoussam_iii,
}


class Command(BaseCommand):
    help = "Importe les cadres logiques PCD (Bafoussam I/II/III) dans la base de données"

    def add_arguments(self, parser):
        parser.add_argument(
            '--reset', action='store_true',
            help='Vide la table PCDObjective avant import',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        if options['reset']:
            count = PCDObjective.objects.count()
            PCDObjective.objects.all().delete()
            self.stdout.write(self.style.WARNING(f"  Table vidée ({count} entrées supprimées)"))

        total = 0
        for commune_nom, filename in FILES.items():
            filepath = os.path.abspath(os.path.join(BASE_DIR, filename))
            if not os.path.exists(filepath):
                self.stdout.write(self.style.ERROR(f"  Fichier introuvable : {filepath}"))
                continue

            self.stdout.write(f"\n→ Lecture {commune_nom} ({filename}) ...")

            try:
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                ws = wb.active
                parser = PARSERS[commune_nom]
                rows = parser(ws)
                wb.close()
            except Exception as e:
                raise CommandError(f"Erreur lecture {filename} : {e}")

            commune_obj = Commune.objects.filter(name__icontains=commune_nom.split()[0]).first()
            if not commune_obj:
                self.stdout.write(self.style.WARNING(
                    f"  Commune '{commune_nom}' non trouvée en base — commune_id sera NULL"
                ))

            to_create = [
                PCDObjective(
                    commune_nom=commune_nom,
                    commune=commune_obj,
                    **row,
                )
                for row in rows
            ]
            PCDObjective.objects.bulk_create(to_create, ignore_conflicts=False)
            self.stdout.write(self.style.SUCCESS(f"  {len(to_create)} objectifs importés"))
            total += len(to_create)

        self.stdout.write(self.style.SUCCESS(f"\nImport terminé — {total} objectifs PCD au total."))
